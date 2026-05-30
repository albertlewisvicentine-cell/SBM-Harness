"""
Excel-based loader for the fault registry.

Public API
----------
load_registry_from_excel(path, sheet) -> Registry

Design constraints
------------------
* Strictly decoupled from renderer/state logic — imports only from models.py
  and the standard library / openpyxl.
* Raises :class:`RegistryLoadError` (a ValueError subclass) for any structural
  problem so callers can catch a single, well-typed exception.
"""

from __future__ import annotations

import pathlib
from typing import Union

import openpyxl

from models import Registry, RegistryEntry

# ---------------------------------------------------------------------------
# Required column names (case-sensitive, must appear in the header row)
# ---------------------------------------------------------------------------
REQUIRED_COLUMNS: tuple[str, ...] = (
    "Code",
    "Category",
    "Severity",
    "DiffusionWeight",
    "RecoveryWeight",
    "ExpectedBaseline",
    "DefaultSubsystem",
    "Description",
)

# Columns that must be cast to float
NUMERIC_COLUMNS: frozenset[str] = frozenset(
    {"DiffusionWeight", "RecoveryWeight", "ExpectedBaseline"}
)


class RegistryLoadError(ValueError):
    """Raised when the registry Excel file cannot be loaded due to invalid content."""


def load_registry_from_excel(
    path: Union[str, pathlib.Path],
    sheet: str = "Registry",
) -> Registry:
    """
    Load a :class:`~models.Registry` from an Excel workbook.

    Parameters
    ----------
    path:
        Path to the ``.xlsx`` workbook.
    sheet:
        Name of the worksheet that contains the registry table.
        Defaults to ``"Registry"``.

    Returns
    -------
    Registry
        A fully-validated :class:`~models.Registry` instance.

    Raises
    ------
    RegistryLoadError
        * If the workbook cannot be opened.
        * If the requested sheet does not exist.
        * If any required column is missing from the header row.
        * If duplicate codes are found.
        * If a numeric cell cannot be cast to ``float``.
    """
    path = pathlib.Path(path)

    # ------------------------------------------------------------------
    # Open workbook
    # ------------------------------------------------------------------
    try:
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    except Exception as exc:
        raise RegistryLoadError(f"Cannot open workbook {path!r}: {exc}") from exc

    if sheet not in wb.sheetnames:
        available = ", ".join(wb.sheetnames)
        raise RegistryLoadError(
            f"Sheet {sheet!r} not found in {path.name!r}. "
            f"Available sheets: {available}"
        )

    ws = wb[sheet]

    # ------------------------------------------------------------------
    # Parse header row
    # ------------------------------------------------------------------
    rows = ws.iter_rows(values_only=True)
    try:
        raw_header = next(rows)
    except StopIteration:
        raise RegistryLoadError(f"Sheet {sheet!r} in {path.name!r} is empty.")

    header: list[str] = [
        (str(cell).strip() if cell is not None else "") for cell in raw_header
    ]

    missing = [col for col in REQUIRED_COLUMNS if col not in header]
    if missing:
        raise RegistryLoadError(
            f"Missing required column(s) in sheet {sheet!r}: {', '.join(missing)}"
        )

    col_index: dict[str, int] = {col: header.index(col) for col in REQUIRED_COLUMNS}

    # ------------------------------------------------------------------
    # Parse data rows
    # ------------------------------------------------------------------
    entries: list[RegistryEntry] = []
    seen_codes: set[str] = set()

    for row_num, raw_row in enumerate(rows, start=2):
        # Skip fully-empty rows
        if all(cell is None or str(cell).strip() == "" for cell in raw_row):
            continue

        def _get(col: str) -> str:
            idx = col_index[col]
            val = raw_row[idx] if idx < len(raw_row) else None
            return str(val).strip() if val is not None else ""

        def _get_float(col: str) -> float:
            raw = _get(col)
            if raw == "":
                raise RegistryLoadError(
                    f"Row {row_num}: numeric column {col!r} is empty."
                )
            try:
                return float(raw)
            except ValueError:
                raise RegistryLoadError(
                    f"Row {row_num}: cannot cast {col!r} value {raw!r} to float."
                )

        code = _get("Code")
        if code == "":
            raise RegistryLoadError(f"Row {row_num}: 'Code' column is empty.")

        if code in seen_codes:
            raise RegistryLoadError(
                f"Duplicate code {code!r} found at row {row_num}."
            )
        seen_codes.add(code)

        entry = RegistryEntry(
            code=code,
            category=_get("Category"),
            severity=_get("Severity"),
            diffusion_weight=_get_float("DiffusionWeight"),
            recovery_weight=_get_float("RecoveryWeight"),
            expected_baseline=_get_float("ExpectedBaseline"),
            default_subsystem=_get("DefaultSubsystem"),
            description=_get("Description"),
        )
        entries.append(entry)

    wb.close()

    return Registry(entries)
