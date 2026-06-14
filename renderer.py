#!/usr/bin/env python3
"""Drift field rendering and throttled replay utilities for spreadsheet visualization."""

from __future__ import annotations

import math
import re
import time
from dataclasses import dataclass
from typing import Callable, Dict, Iterable, List, Sequence, Tuple

try:
    from openpyxl.styles import PatternFill
except ImportError:  # pragma: no cover - optional dependency
    PatternFill = None


@dataclass(frozen=True)
class TopologyCell:
    """Mapping for one state index to one spreadsheet coordinate."""

    index: int
    row: int
    col: int
    coordinate: str


class DriftRenderer:
    """Render a drift field as a spreadsheet heatmap and replay synthetic events."""

    def __init__(self, worksheet, topology_shape: Tuple[int, int], anchor_cell: str = "A1"):
        self.worksheet = worksheet
        self.rows, self.cols = topology_shape
        if self.rows <= 0 or self.cols <= 0:
            raise ValueError("topology_shape must have positive row and column counts")

        self.start_row, self.start_col = _parse_cell(anchor_cell)
        self.topology: Dict[int, TopologyCell] = self._build_topology()
        self._state_field = [0.0] * (self.rows * self.cols)

    def _build_topology(self) -> Dict[int, TopologyCell]:
        topology = {}
        for index in range(self.rows * self.cols):
            rel_row = index // self.cols
            rel_col = index % self.cols
            row = self.start_row + rel_row
            col = self.start_col + rel_col
            topology[index] = TopologyCell(
                index=index,
                row=row,
                col=col,
                coordinate=_to_coordinate(row, col),
            )
        return topology

    def map_index_to_cell(self, index: int) -> str:
        """Return spreadsheet coordinate wired to a linear state index."""
        if index not in self.topology:
            raise IndexError(f"State index out of range: {index}")
        return self.topology[index].coordinate

    def render(self, drift_field: Sequence[float]) -> None:
        """Render a drift field as a heatmap where drift magnitude drives cell color."""
        values = list(drift_field)
        expected = self.rows * self.cols
        if len(values) != expected:
            raise ValueError(f"drift_field length must be {expected}, got {len(values)}")

        self._state_field = [float(v) for v in values]
        for index, value in enumerate(self._state_field):
            cell_ref = self.map_index_to_cell(index)
            rgb = _drift_to_rgb(value)
            self._paint_cell(cell_ref, rgb, value)

    def replay(
        self,
        event_stream: Iterable,
        fps: float = 4.0,
        sleep_fn: Callable[[float], None] = time.sleep,
        clock_fn: Callable[[], float] = time.monotonic,
    ) -> int:
        """Replay a synthetic event stream with a predictable throttled frame cadence."""
        if fps < 0:
            raise ValueError("fps must be non-negative")

        frame_interval = 1.0 / fps if fps else 0.0
        next_frame_time = clock_fn()
        rendered_frames = 0

        for event in event_stream:
            drift_field = self._event_to_field(event)
            self.render(drift_field)
            rendered_frames += 1

            if frame_interval:
                next_frame_time += frame_interval
                delay = next_frame_time - clock_fn()
                if delay > 0:
                    sleep_fn(delay)
                else:
                    next_frame_time = clock_fn()

        return rendered_frames

    def _event_to_field(self, event) -> List[float]:
        if isinstance(event, (list, tuple)):
            if len(event) != len(self._state_field):
                raise ValueError("Event drift field length does not match topology")
            self._state_field = [float(v) for v in event]
            return list(self._state_field)

        if isinstance(event, dict):
            if "drift_field" in event:
                return self._event_to_field(event["drift_field"])

            updates = event.get("updates")
            if updates is None and "index" in event:
                updates = [event]

            if updates is not None:
                state = list(self._state_field)
                for update in updates:
                    idx = int(update["index"])
                    if idx < 0 or idx >= len(state):
                        raise IndexError(f"State index out of range: {idx}")
                    state[idx] = float(update.get("drift", 0.0))
                self._state_field = state
                return list(self._state_field)

        raise ValueError("Unsupported event format for replay")

    def _paint_cell(self, cell_ref: str, rgb: Tuple[int, int, int], value: float) -> None:
        if hasattr(self.worksheet, "range"):  # xlwings-like sheet
            cell = self.worksheet.range(cell_ref)
            cell.value = value
            cell.color = rgb
            return

        cell = self.worksheet[cell_ref]
        cell.value = value
        if PatternFill is not None:
            hex_rgb = "{:02X}{:02X}{:02X}".format(*rgb)
            cell.fill = PatternFill(fill_type="solid", start_color=hex_rgb, end_color=hex_rgb)
        else:  # pragma: no cover - optional dependency fallback
            cell.fill = {"rgb": rgb}


class Renderer(DriftRenderer):
    """Contract-compatible renderer alias."""


def generate_fault_bloom_events(
    topology_shape: Tuple[int, int],
    epicenter_index: int,
    frames: int = 8,
    peak_drift: float = 1.0,
    decay: float = 0.72,
) -> List[dict]:
    """Generate a synthetic 'fault bloom' stream that spreads outward across topology."""
    rows, cols = topology_shape
    size = rows * cols
    if size <= 0:
        raise ValueError("topology_shape must have positive size")
    if epicenter_index < 0 or epicenter_index >= size:
        raise IndexError("epicenter_index out of range")

    center_row = epicenter_index // cols
    center_col = epicenter_index % cols

    events = []
    for frame in range(frames):
        drift_field = []
        frame_peak = peak_drift * (decay ** frame)
        for index in range(size):
            row = index // cols
            col = index % cols
            distance = math.hypot(row - center_row, col - center_col)
            drift = frame_peak * math.exp(-distance)
            drift_field.append(float(drift))
        events.append({"drift_field": drift_field})

    return events


def _drift_to_rgb(value: float) -> Tuple[int, int, int]:
    value = max(-1.0, min(1.0, float(value)))
    if value >= 0:
        fade = int(round(255 * (1.0 - value)))
        return 255, fade, fade

    fade = int(round(255 * (1.0 + value)))
    return fade, fade, 255


def _parse_cell(cell_ref: str) -> Tuple[int, int]:
    match = re.fullmatch(r"([A-Za-z]+)(\d+)", cell_ref)
    if not match:
        raise ValueError(f"Invalid anchor cell reference: {cell_ref}")

    col_letters, row_text = match.groups()
    row = int(row_text)
    col = 0
    for char in col_letters.upper():
        col = col * 26 + (ord(char) - ord("A") + 1)

    return row, col


def _to_coordinate(row: int, col: int) -> str:
    if row <= 0 or col <= 0:
        raise ValueError("Row and column must be positive")

    letters = []
    cur = col
    while cur:
        cur, rem = divmod(cur - 1, 26)
        letters.append(chr(ord("A") + rem))
    return "".join(reversed(letters)) + str(row)
